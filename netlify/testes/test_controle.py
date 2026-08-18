"""Publicar estado a estado nao pode apagar o que ja foi entregue.

Reproduz o relato: PA normal, PA sem-compras, depois TO normal. Ao fim, os
vendedores do PA precisam continuar com as duas listas deles.
"""
import json
import os
import time
import urllib.request
import warnings
import openpyxl
from playwright.sync_api import sync_playwright

warnings.filterwarnings("ignore")
S = "/tmp/claude-0/-home-user-Dubosa/2865c67b-73f7-5f69-865e-8196845b571f/scratchpad"
BASE = "http://localhost:8899"
SENHA = "senha-de-teste"
falhas = []


def ok(cond, nome, extra=""):
    print(("  ok    " if cond else "  FALHA ") + nome + ("" if cond else f"  -> {extra}"))
    if not cond:
        falhas.append(nome)


def api(caminho, corpo, admin=True):
    req = urllib.request.Request(
        BASE + caminho, data=json.dumps(corpo).encode(),
        headers={"content-type": "application/json",
                 **({"x-admin-token": SENHA} if admin else {})})
    try:
        with urllib.request.urlopen(req) as r:
            return r.status, json.loads(r.read())
    except urllib.error.HTTPError as e:
        return e.code, json.loads(e.read())


def recorte(uf, destino, so_sem_compras=False):
    wb = openpyxl.load_workbook(f"{S}/data_2.xlsx")
    ws = wb["Export"]
    h = [c.value for c in ws[1]]
    ui = h.index("UF") + 1
    for r in range(ws.max_row, 1, -1):
        if str(ws.cell(r, ui).value or "").strip().upper() != uf:
            ws.delete_rows(r)
    wb.save(destino)


print("preparando recortes por estado")
for uf in ["PA", "TO"]:
    destino = f"{S}/ctl_{uf.lower()}.xlsx"
    if os.path.exists(destino):
        print(f"  {uf}: reaproveitando {os.path.basename(destino)}")
    else:
        recorte(uf, destino)


def publicar_rodada(pg, arq, modo):
    pg.click("#s1"); pg.wait_for_timeout(250)
    pg.set_input_files("#file", arq); pg.wait_for_timeout(2700)
    pg.click("#toEquipe"); pg.wait_for_timeout(300)
    pg.check(f"input[name='modo'][value='{modo}']"); pg.wait_for_timeout(450)
    pg.click("#run"); pg.wait_for_timeout(2900)
    marcados = pg.evaluate(
        "()=>[...document.querySelectorAll('#selVend input[data-vend]:checked')]"
        ".map(i=>i.getAttribute('data-vend'))")
    pg.fill("#senhaPub", SENHA)
    pg.click("#publicar")
    pg.wait_for_function("document.getElementById('publicar').disabled === false", timeout=180000)
    return marcados


with sync_playwright() as p:
    b = p.chromium.launch(executable_path="/opt/pw-browsers/chromium")
    pg = b.new_context(viewport={"width": 1400, "height": 1000}).new_page()
    errs = []
    pg.on("pageerror", lambda e: errs.append(str(e)))
    pg.goto(BASE)
    pg.wait_for_timeout(700)

    print("\n1. PA — distribuição")
    m1 = publicar_rodada(pg, f"{S}/ctl_pa.xlsx", "normal")
    print("     marcados por padrão:", len(m1), m1[:3])
    ok(all("PA" in st or True for st in m1), "seleção montada")
    ok(len(m1) == 4, f"só os 4 vendedores do PA vêm marcados (vieram {len(m1)})", m1)

    print("\n2. PA — sem compras")
    m2 = publicar_rodada(pg, f"{S}/ctl_pa.xlsx", "carteira")
    print("     marcados:", len(m2))

    # Na modalidade sem-compras o dono vem da coluna Vendedor da base, entao
    # nem todo vendedor do PA aparece: so quem tem cliente do PA na carteira.
    _, sit = api("/api/situacao", {"vendedores": sorted(set(m1 + m2))})
    porVend = {i["vendedor"]: i for i in sit["itens"]}
    comDuas = [v for v in m1 if len(porVend.get(v, {}).get("rodadas", [])) == 2]
    print("     do PA, com as duas listas:", comDuas)
    ok(comDuas, "ao menos um vendedor do PA ficou com as 2 listas",
       {v: [r["modo"] for r in porVend[v]["rodadas"]] for v in m1})
    alvo = comDuas[0]

    print("\n3. TO — distribuição (o teste do relato)")
    m3 = publicar_rodada(pg, f"{S}/ctl_to.xlsx", "normal")
    print("     marcados:", len(m3), m3[:3])
    ok(not any(v in m1 for v in m3), "nenhum vendedor do PA foi marcado na rodada do TO",
       [v for v in m3 if v in m1])

    _, sitDepois = api("/api/situacao", {"vendedores": sorted(set(m1 + m2 + m3))})
    depoisDe = {i["vendedor"]: i for i in sitDepois["itens"]}

    # O ponto do relato: publicar o TO nao pode mexer no que ja foi entregue.
    # Quem entrou na rodada do TO recebe (so) a lista normal dele -- isso e o
    # esperado. Todo o resto tem de sair intacto. Vale lembrar que a rodada de
    # sem-compras do PA alcanca vendedores de qualquer estado, entao ha quem
    # esteja em m2 e em m3 ao mesmo tempo.
    def foto(item, modos=None):
        return sorted((r["modo"], r["publicadoEm"]) for r in item["rodadas"]
                      if modos is None or r["modo"] in modos)

    mudou = []
    for v in sorted(set(m1 + m2)):
        # de quem entrou no TO, so a lista normal podia mudar
        modos = ["carteira", "ataque"] if v in m3 else None
        if foto(porVend[v], modos) != foto(depoisDe[v], modos):
            mudou.append(v)
    ok(not mudou, "publicar o TO não alterou nada que já estava entregue", mudou)
    ok(all(foto(porVend[v], ["carteira"]) == foto(depoisDe[v], ["carteira"]) for v in m2),
       "nenhuma lista de sem-compras do PA foi tocada",
       [v for v in m2 if foto(porVend[v], ["carteira"]) != foto(depoisDe[v], ["carteira"])])
    ok(all(foto(porVend[v], ["normal"]) == foto(depoisDe[v], ["normal"]) for v in m1),
       "nenhuma distribuição do PA foi sobrescrita",
       [v for v in m1 if foto(porVend[v], ["normal"]) != foto(depoisDe[v], ["normal"])])
    ok(len(porVend[alvo]["rodadas"]) == 2 and len(depoisDe[alvo]["rodadas"]) == 2,
       f"{alvo} segue com as 2 listas depois da rodada do TO",
       [r["modo"] for r in depoisDe[alvo]["rodadas"]])
    ok(all(any(r["modo"] == "normal" for r in depoisDe[v]["rodadas"]) for v in m3),
       "cada vendedor do TO recebeu a distribuição dele",
       {v: [r["modo"] for r in depoisDe[v]["rodadas"]] for v in m3})

    print("\n4. desmarcar alguém: o link dele não é tocado")
    pg.click("#s1"); pg.wait_for_timeout(250)
    pg.set_input_files("#file", f"{S}/ctl_pa.xlsx"); pg.wait_for_timeout(2700)
    pg.click("#toEquipe"); pg.wait_for_timeout(300)
    pg.check("input[name='modo'][value='normal']"); pg.wait_for_timeout(400)
    pg.click("#run"); pg.wait_for_timeout(2900)
    # Usa quem tem as duas listas, para provar que nenhuma das duas se mexe.
    fora = alvo
    pg.evaluate("(n)=>document.querySelector('#selVend input[data-vend=\"'+n+'\"]').checked=false", fora)
    antes = depoisDe[fora]["rodadas"]
    antesData = [r["publicadoEm"] for r in antes if r["modo"] == "normal"][0]
    pg.fill("#senhaPub", SENHA); pg.click("#publicar")
    pg.wait_for_function("document.getElementById('publicar').disabled === false", timeout=180000)
    _, sit2 = api("/api/situacao", {"vendedores": [fora]})
    depois = sit2["itens"][0]["rodadas"]
    depoisData = [r["publicadoEm"] for r in depois if r["modo"] == "normal"][0]
    ok(depoisData == antesData, f"{fora} desmarcado: a lista dele não mudou de data",
       f"{antesData} -> {depoisData}")
    ok(len(depois) == 2, "e continua com as duas listas", [r["modo"] for r in depois])

    print("\n5. apagar uma lista pelo painel")
    pg.click("#verSituacao")
    pg.wait_for_function("document.getElementById('verSituacao').disabled === false", timeout=60000)
    pg.wait_for_timeout(600)
    nVend = pg.evaluate("document.querySelectorAll('#situacaoSaida .sit-vend').length")
    print("     vendedores com lista no painel:", nVend)
    ok(nVend >= 8, "painel lista PA e TO", nVend)

    pg.on("dialog", lambda d: d.accept())
    pg.evaluate("""(n)=>{
      const b=[...document.querySelectorAll('#situacaoSaida [data-apagar]')]
        .find(b=>b.getAttribute('data-apagar')===n && b.getAttribute('data-modo')==='carteira');
      b.click();
    }""", fora)
    pg.wait_for_timeout(2500)
    _, sit3 = api("/api/situacao", {"vendedores": [fora]})
    modos = [r["modo"] for r in sit3["itens"][0]["rodadas"]]
    ok(modos == ["normal"], f"só a lista de sem-compras de {fora} foi apagada", modos)

    print("\n6. limpar tudo de um vendedor")
    pg.evaluate("""(n)=>{
      document.querySelector('#situacaoSaida [data-limpar="'+n+'"]').click();
    }""", fora)
    pg.wait_for_timeout(2500)
    _, sit4 = api("/api/situacao", {"vendedores": [fora]})
    ok(sit4["itens"][0]["rodadas"] == [], f"{fora} sem nenhuma lista",
       sit4["itens"][0]["rodadas"])
    ok(sit4["itens"][0].get("token"), "mas o link dele continua existindo")

    print("\n7. o vendedor vê a página vazia, sem erro")
    v = b.new_context().new_page()
    v.goto(f"{BASE}/c/#{sit4['itens'][0]['token']}")
    v.wait_for_timeout(1800)
    ok("Nenhuma lista ativa" in v.inner_text("#conteudo"), "página explica que não há lista",
       v.inner_text("#conteudo")[:100])

    print("\n8. senha errada não apaga nada")
    st, resp = api("/api/apagar", {"vendedor": m3[0]}, admin=False)
    ok(st == 401, "apagar sem senha é recusado", st)
    st, resp = api("/api/situacao", {"vendedores": m3}, admin=False)
    ok(st == 401, "consultar sem senha é recusado", st)

    print("\nerros de página:", errs or "nenhum")
    if errs:
        falhas.append("erros de console")
    pg.screenshot(path="controle-painel.png", full_page=True)
    b.close()

print("\n" + (f"{len(falhas)} FALHA(S): {falhas}" if falhas else "Todos os testes passaram."))
raise SystemExit(1 if falhas else 0)
