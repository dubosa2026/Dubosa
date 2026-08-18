"""Sem compras no mes: as duas regras de destino do cliente.

Regra 1 - o cliente ja e de um vendedor DA EQUIPE: fica com ele.
Regra 2 - o cliente esta com vendedor de FORA da equipe: passa para um
          vendedor da equipe naquele estado.

Ninguem de fora da equipe pode aparecer na lista de publicacao.
"""
import warnings
import openpyxl
from playwright.sync_api import sync_playwright

warnings.filterwarnings("ignore")
S = "/tmp/claude-0/-home-user-Dubosa/2865c67b-73f7-5f69-865e-8196845b571f/scratchpad"
APP = "/home/user/Dubosa/app/dist/index.html"
falhas = []


def ok(cond, nome, extra=""):
    print(("  ok    " if cond else "  FALHA ") + nome + ("" if cond else f"  -> {extra}"))
    if not cond:
        falhas.append(nome)


def base_pa(destino):
    """Recorta o PA da base real e devolve quem aparece na coluna Vendedor."""
    wb = openpyxl.load_workbook(f"{S}/data_2.xlsx")
    ws = wb["Export"]
    h = [c.value for c in ws[1]]
    ui, vi = h.index("UF") + 1, h.index("Vendedor") + 1
    donos = {}
    for r in range(ws.max_row, 1, -1):
        if str(ws.cell(r, ui).value or "").strip().upper() != "PA":
            ws.delete_rows(r)
            continue
        donos[str(ws.cell(r, vi).value or "").strip()] = True
    wb.save(destino)
    return sorted(d for d in donos if d)


donos = base_pa(f"{S}/sc_pa.xlsx")
print(f"base do PA: {len(donos)} nomes na coluna Vendedor")


def rodar_carteira(pg, arq):
    pg.click("#s1")
    pg.wait_for_timeout(250)
    pg.set_input_files("#file", arq)
    pg.wait_for_timeout(2700)
    pg.click("#toEquipe")
    pg.wait_for_timeout(300)
    pg.check("input[name='modo'][value='carteira']")
    pg.wait_for_timeout(500)
    pg.click("#run")
    pg.wait_for_timeout(2900)
    return pg.evaluate("""() => ({
      resumo: state.result.resumo.map(v => ({
        vendedor: v.vendedor, uf: v.uf, naEquipe: v.naEquipe,
        qtde: v.qtde, mantidos: v.mantidos, repassados: v.repassados
      })),
      repassados: state.result.repassados,
      semEquipeNaUf: state.result.semEquipeNaUf.length,
      semDono: state.result.semDono.length,
      atribuidos: state.result.atribuidos.length,
      grafias: state.result.grafiasAceitas
    })""")


with sync_playwright() as p:
    b = p.chromium.launch(executable_path="/opt/pw-browsers/chromium")
    pg = b.new_context(viewport={"width": 1400, "height": 1000}).new_page()
    errs = []
    pg.on("pageerror", lambda e: errs.append(str(e)))
    pg.on("console", lambda m: errs.append(m.text) if m.type == "error" else None)
    pg.goto("file://" + APP)
    pg.wait_for_timeout(700)

    equipe = pg.evaluate("state.equipe.map(v => v.vendedor).filter(Boolean)")
    print(f"equipe cadastrada: {len(equipe)} vendedores")

    print("\n1. ninguem de fora da equipe entra na lista")
    r = rodar_carteira(pg, f"{S}/sc_pa.xlsx")
    fora = [v["vendedor"] for v in r["resumo"] if v["vendedor"] not in equipe]
    print(f"     vendedores na lista: {len(r['resumo'])} (base tinha {len(donos)} donos)")
    ok(not fora, "nenhum vendedor de fora da equipe na lista", fora)
    ok(all(v["naEquipe"] for v in r["resumo"]), "todos marcados como da equipe")
    ok(len(r["resumo"]) < len(donos),
       "a lista encolheu: os de fora foram absorvidos pela equipe",
       f"{len(r['resumo'])} vs {len(donos)}")

    print("\n2. os clientes de fora foram repassados, nao perdidos")
    print(f"     repassados: {r['repassados']} | sem equipe na UF: {r['semEquipeNaUf']}")
    ok(r["repassados"] > 0, "houve repasse", r["repassados"])
    soma = sum(v["repassados"] for v in r["resumo"])
    ok(soma == r["repassados"], "a soma por vendedor bate com o total",
       f"{soma} vs {r['repassados']}")
    ok(r["semEquipeNaUf"] == 0, "no PA a equipe existe, entao nada ficou sem destino",
       r["semEquipeNaUf"])

    print("\n3. quem recebeu repasse e da equipe do PA")
    equipePa = pg.evaluate("state.equipe.filter(v=>v.uf==='PA'&&v.vendedor).map(v=>v.vendedor)")
    receberam = [v["vendedor"] for v in r["resumo"] if v["repassados"]]
    print("     equipe do PA:", equipePa)
    print("     receberam repasse:", receberam)
    ok(set(receberam) <= set(equipePa), "o repasse ficou na equipe do estado",
       [v for v in receberam if v not in equipePa])

    print("\n4. o rodizio dividiu por igual entre eles")
    qts = sorted(v["repassados"] for v in r["resumo"] if v["repassados"])
    print("     repassados por vendedor:", qts)
    ok(not qts or qts[-1] - qts[0] <= 1, "diferenca maxima de 1 cliente", qts)

    print("\n5. quem ja era da equipe manteve os proprios clientes")
    mantidos = [v for v in r["resumo"] if v["mantidos"]]
    print(f"     {len(mantidos)} vendedores com cliente proprio")
    ok(mantidos, "houve cliente mantido com o dono original")
    ok(all(v["mantidos"] + v["repassados"] == v["qtde"] for v in r["resumo"]),
       "mantidos + repassados = total de cada vendedor",
       [(v["vendedor"], v["mantidos"], v["repassados"], v["qtde"]) for v in r["resumo"]])

    print("\n6. a coluna Vendedor da linha repassada mostra o novo dono")
    checa = pg.evaluate("""() => {
      const col = state.result.colVendedorBase;
      const rep = state.result.atribuidos.filter(r => r.__origemCarteira === 'repassado');
      return {
        total: rep.length,
        certos: rep.filter(r => r[col] === r.__vendedor).length,
        vazaram: rep.filter(r => r[col] === r.__donoAnterior).length,
        temAnterior: rep.filter(r => r.__donoAnterior).length
      };
    }""")
    print("     ", checa)
    ok(checa["certos"] == checa["total"], "toda linha repassada traz o novo responsavel", checa)
    ok(checa["vazaram"] == 0, "o nome do vendedor de fora nao vai na lista publicada", checa)
    ok(checa["temAnterior"] == checa["total"], "o dono anterior fica guardado para o gestor", checa)

    print("\n7. a tela explica o que aconteceu")
    sub = pg.inner_text("#resSub")
    print("     ", sub[:200])
    ok("de fora da equipe" in sub, "o texto cita o repasse", sub[:120])
    tags = pg.evaluate("document.querySelectorAll('#vendList .tag-repasse').length")
    ok(tags == len(receberam), "etiqueta de repasse em quem recebeu", f"{tags} vs {len(receberam)}")

    print("\n8. estado sem vendedor cadastrado: o cliente nao some")
    # SP nao tem ninguem na equipe. Roda a base inteira com SP no filtro.
    pg.click("#s1"); pg.wait_for_timeout(250)
    pg.set_input_files("#file", f"{S}/data_3.xlsx"); pg.wait_for_timeout(3000)
    pg.click("#toEquipe"); pg.wait_for_timeout(300)
    pg.check("input[name='modo'][value='carteira']"); pg.wait_for_timeout(500)
    marcou = pg.evaluate("""() => {
      const NORTE = ['AC','AM','AP','PA','RO','RR','TO'];
      const ufs = [...document.querySelectorAll('#ufFiltro input[data-uf]')];
      ufs.forEach(i => { i.checked = false; });
      const alvo = ufs.filter(i => !NORTE.includes(i.getAttribute('data-uf')));
      if (!alvo.length) return null;
      alvo[0].checked = true;
      alvo[0].dispatchEvent(new Event('change', {bubbles:true}));
      return alvo[0].getAttribute('data-uf');
    }""")
    print("     estado escolhido (sem equipe):", marcou)
    if marcou:
        pg.click("#run"); pg.wait_for_timeout(2900)
        f = pg.evaluate("""() => ({
          semEquipe: state.result.semEquipeNaUf.length,
          resumo: state.result.resumo.length,
          repassados: state.result.repassados
        })""")
        print("     ", f)
        ok(f["semEquipe"] > 0 or f["resumo"] > 0,
           f"os clientes de {marcou} apareceram em algum lugar", f)
        ok(f["repassados"] == 0, "sem equipe no estado, nada foi repassado", f)
        texto = pg.inner_text("#buckets")
        ok("Estado sem ninguém da sua equipe" in texto,
           "o relatorio mostra o grupo sem destino", texto[:150])
        pg.screenshot(path="sem-compras-sem-equipe.png", full_page=True)
    else:
        print("     (a base nao tem estado fora do Norte; passo pulado)")

    ov = pg.evaluate("document.documentElement.scrollWidth > document.documentElement.clientWidth + 1")
    ok(not ov, "sem overflow horizontal")
    print("\nerros de pagina:", errs or "nenhum")
    if errs:
        falhas.append("erros de console")
    b.close()

print("\n" + (f"{len(falhas)} FALHA(S): {falhas}" if falhas else "Todos os testes passaram."))
raise SystemExit(1 if falhas else 0)
