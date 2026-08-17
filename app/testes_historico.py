"""Correcoes das analises + historico por estado.

Gera bases sinteticas derivadas da real para simular varias rodadas com
taxas de conversao conhecidas, e confere se a serie por estado bate.
"""
import random
import openpyxl
import warnings
from playwright.sync_api import sync_playwright

warnings.filterwarnings("ignore")
S = "/tmp/claude-0/-home-user-Dubosa/2865c67b-73f7-5f69-865e-8196845b571f/scratchpad"

falhas = []


def ok(cond, nome, extra=""):
    print(("  ok    " if cond else "  FALHA ") + nome + ("" if cond else f"  -> {extra}"))
    if not cond:
        falhas.append(nome)


def gerar_rodada(origem, destino, fracao, semente):
    """Copia a base convertendo `fracao` dos parados em 'Ativo 30 dias'."""
    random.seed(semente)
    wb = openpyxl.load_workbook(origem)
    ws = wb["Export"]
    ci = [c.value for c in ws[1]].index("Categoria") + 1
    virou = 0
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(r, ci).value
        if cat in ("Inativo", "Sem Compras", "Ativo 60 dias") and random.random() < fracao:
            ws.cell(r, ci).value = "Ativo 30 dias"
            virou += 1
    wb.save(destino)
    return virou


# 3 rodadas com conversao crescente, todas a partir da base original
for i, frac in enumerate([0.05, 0.12, 0.20], start=1):
    n = gerar_rodada(f"{S}/data_2.xlsx", f"{S}/hist_{i}.xlsx", frac, 100 + i)
    print(f"base hist_{i}.xlsx gerada: {n} clientes viraram Ativo 30 dias ({frac:.0%})")

# O arquivo gerado ja e um documento HTML completo: envelopa-lo dentro de
# outro <html> faz o navegador descartar parte do conteudo. Abre direto.
APP = "/home/user/Dubosa/app/dist/index.html"


def rodar(pg, arq):
    pg.click("#s1"); pg.wait_for_timeout(250)
    pg.set_input_files("#file", arq); pg.wait_for_timeout(3000)
    pg.click("#toEquipe"); pg.wait_for_timeout(300)
    pg.click("#run"); pg.wait_for_timeout(3200)


with sync_playwright() as p:
    b = p.chromium.launch(executable_path="/opt/pw-browsers/chromium")
    pg = b.new_context(viewport={"width": 1400, "height": 1000}).new_page()
    errs = []
    pg.on("pageerror", lambda e: errs.append(str(e)))
    pg.on("console", lambda m: errs.append(m.text) if m.type == "error" else None)
    pg.goto("file://" + APP)
    pg.wait_for_timeout(700)

    print("\n1. primeira rodada — nada com que comparar")
    rodar(pg, f"{S}/data_2.xlsx")
    pg.click("#s4"); pg.wait_for_timeout(600)
    # O funil agora sempre existe; o que muda e quais estados entraram na
    # conta. Na primeira rodada, nenhum tem referencia ainda.
    ok(pg.evaluate("state.funil.ufsComparadas.length") == 0, "nenhum estado comparado")
    ok(pg.evaluate("state.funil.ufsSemReferencia.length") == 7, "os 7 estados sem referência")
    ok("passam a contar" in pg.inner_text("#perfSub"), "explica que começam a contar agora")
    ok(pg.evaluate("lerHistorico().length") == 0, "histórico ainda vazio")

    print("\n2. mesma base de novo — não é um período")
    rodar(pg, f"{S}/data_2.xlsx")
    pg.click("#s4"); pg.wait_for_timeout(600)
    ok(pg.evaluate("state.funil.ufsMesmaBase.length") == 7,
       "detectou a mesma base nos 7 estados",
       pg.evaluate("state.funil.ufsMesmaBase"))
    ok("mesma base" in pg.inner_text("#perfSub"), "explica na tela", pg.inner_text("#perfSub")[:90])
    ok(pg.evaluate("lerHistorico().length") == 0, "não gravou no histórico")
    titulos = pg.evaluate("state.insights.map(i=>i.titulo)")
    ok(not any("Aproveitamento" in t for t in titulos), "não inventa aproveitamento", titulos)

    print("\n3. rodadas com conversão de verdade")
    for i, esperado in enumerate([0.05, 0.12, 0.20], start=1):
        rodar(pg, f"{S}/hist_{i}.xlsx")
        f = pg.evaluate("state.funil && {t:state.funil.taxaGeral,c:state.funil.confiavel,"
                        "conv:state.funil.totalConversoes,cart:state.funil.totalCarteira,"
                        "perd:state.funil.perdidosDeVista}")
        soma_ok = f["conv"] + pg.evaluate("state.funil.aindaAbertos") + f["perd"] == f["cart"]
        print(f"     rodada {i}: taxa {f['t']*100:.1f}% (esperado ~{esperado*100:.0f}%), "
              f"confiável={f['c']}, sumiram={f['perd']}")
        ok(f["c"], f"rodada {i} considerada confiável")
        ok(soma_ok, f"rodada {i}: converteu + aberto + sumiu = carteira")
        ok(abs(f["t"] - esperado) < 0.05, f"rodada {i}: taxa próxima do esperado")

    hist = pg.evaluate("lerHistorico()")
    ok(len(hist) == 3, f"histórico com 3 rodadas (tem {len(hist)})")

    print("\n4. coerência das análises")
    ins = pg.evaluate("state.insights.map(i=>({n:i.nivel,t:i.titulo,x:i.texto}))")
    for i in ins:
        print(f"     [{i['n']}] {i['t']}")
    destaque = [i for i in ins if i["t"].startswith("Destaque")]
    zerados = [i for i in ins if "sem nenhuma conversão" in i["t"]]
    if destaque:
        nome = destaque[0]["t"].replace("Destaque: ", "")
        ok("Converteu 0 " not in destaque[0]["x"], "destaque tem conversão > 0", destaque[0]["x"][:70])
        ok(not zerados or nome not in zerados[0]["x"],
           "destaque não aparece na lista de quem zerou", nome)
    ok(not any("R$ 0" in i["x"] for i in ins), "não fala de R$ 0",
       [i["x"][:60] for i in ins if "R$ 0" in i["x"]])

    print("\n5. histórico por estado")
    pg.click("#s4"); pg.wait_for_timeout(700)
    ufs = pg.evaluate("()=>[...document.querySelectorAll('#histUfs input')].map(i=>i.value)")
    print("     seletor:", ufs)
    ok(ufs[0] == "" and len(ufs) == 8, "equipe toda + 7 estados do Norte", ufs)

    # confere PA contra o dado bruto do historico
    pg.evaluate("()=>document.querySelector('#histUfs input[value=PA]').click()")
    pg.wait_for_timeout(600)
    esperado_pa = [
        (h["ufs"] and [u for u in h["ufs"] if u["uf"] == "PA"][0]) for h in hist
    ]
    na_tela = pg.evaluate("""()=>[...document.querySelectorAll('#histConteudo tbody tr')]
        .map(tr=>[...tr.children].map(td=>td.textContent))""")
    print("     tabela PA (mais recente primeiro):")
    for linha in na_tela:
        print("      ", linha)
    ok(len(na_tela) == 3, "3 linhas para PA", len(na_tela))
    conv_tela = [int(l[3]) for l in na_tela][::-1]
    conv_hist = [u["conversoes"] for u in esperado_pa]
    ok(conv_tela == conv_hist, "reativados por rodada batem com o histórico",
       f"tela={conv_tela} hist={conv_hist}")
    barras = pg.evaluate("document.querySelectorAll('#histConteudo .hbar').length")
    ok(barras == 3, "3 barras no gráfico", barras)
    tend = pg.inner_text("#histConteudo")
    ok("▲" in tend or "▼" in tend or "=" in tend, "mostra a tendência")
    pg.screenshot(path="hist-pa.png", full_page=True)

    pg.evaluate("()=>document.querySelector('#histUfs input[value=\\'\\']').click()")
    pg.wait_for_timeout(600)
    todos = pg.evaluate("document.querySelectorAll('#histConteudo tbody tr').length")
    ok(todos == 3, "equipe toda também com 3 rodadas", todos)
    pg.screenshot(path="hist-todos.png", full_page=True)

    print("\n6. bases que não se cruzam")
    rodar(pg, f"{S}/base_mes.xlsx")   # base de PA, modo normal: clientes diferentes
    f = pg.evaluate("state.funil && {conf:state.funil.confiavel,"
                    "naoConf:(state.funil.ufsNaoConfiaveis||[]).map(u=>u.uf),"
                    "perd:state.funil.perdidosDeVista,cart:state.funil.totalCarteira}")
    pg.click("#s4"); pg.wait_for_timeout(600)
    sub = pg.inner_text("#perfSub")
    print("     funil:", f)
    print("     aviso:", sub[:130])
    ok(f and not f["conf"], "marcou a comparação como não confiável")
    ok("ficou de fora porque" in sub, "explica por estado em vez de mostrar 0%", sub[:120])
    ok(pg.evaluate("lerHistorico().length") == 3, "não sujou o histórico")
    titulos = pg.evaluate("state.insights.map(i=>i.titulo)")
    ok(not any("Aproveitamento" in t for t in titulos), "sem cartão de aproveitamento", titulos)
    ok(not any(t.startswith("Destaque") for t in titulos), "sem destaque falso", titulos)
    pg.screenshot(path="hist-nao-cruza.png", full_page=True)

    ov = pg.evaluate("document.documentElement.scrollWidth > document.documentElement.clientWidth + 1")
    ok(not ov, "sem overflow horizontal")
    print("\nerros de página:", errs or "nenhum")
    if errs:
        falhas.append("erros de console")
    b.close()

print("\n" + (f"{len(falhas)} FALHA(S): {falhas}" if falhas else "Todos os testes passaram."))
raise SystemExit(1 if falhas else 0)
