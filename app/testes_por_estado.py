"""Referencia por estado: o cenario real do gestor.

Ele carrega a base do PA, distribui; depois a do TO, distribui. Cada estado
tem de ser comparado com a PROPRIA rodada anterior -- rodar o TO nao pode
apagar nem contaminar a referencia do PA.
"""
import random
import warnings
import openpyxl
from playwright.sync_api import sync_playwright

warnings.filterwarnings("ignore")
S = "/tmp/claude-0/-home-user-Dubosa/2865c67b-73f7-5f69-865e-8196845b571f/scratchpad"
APP = "/home/user/Dubosa/app/dist/index.html"
falhas = []


def ok(cond, nome, extra=""):
    print(("  ok    " if cond else "  FALHA ") + nome + ("" if cond else f"  -> {extra}"))
    if not cond:
        falhas.append(nome)


def base_do_estado(uf, destino, fracao_convertida=0.0, semente=1):
    """Recorta a base real num estado só, convertendo parte dos parados."""
    random.seed(semente)
    wb = openpyxl.load_workbook(f"{S}/data_2.xlsx")
    ws = wb["Export"]
    h = [c.value for c in ws[1]]
    ui, ci = h.index("UF") + 1, h.index("Categoria") + 1
    virou = 0
    for r in range(ws.max_row, 1, -1):
        if str(ws.cell(r, ui).value or "").strip().upper() != uf:
            ws.delete_rows(r)
            continue
        if (ws.cell(r, ci).value in ("Inativo", "Sem Compras", "Ativo 60 dias")
                and random.random() < fracao_convertida):
            ws.cell(r, ci).value = "Ativo 30 dias"
            virou += 1
    wb.save(destino)
    return ws.max_row - 1, virou


print("preparando bases por estado")
for uf in ["PA", "TO"]:
    n, _ = base_do_estado(uf, f"{S}/{uf.lower()}_1.xlsx", 0.0, 1)
    n2, v = base_do_estado(uf, f"{S}/{uf.lower()}_2.xlsx", 0.15, 2)
    print(f"  {uf}: {n} linhas; 2ª rodada com {v} convertidos")


def rodar(pg, arq):
    pg.click("#s1"); pg.wait_for_timeout(250)
    pg.set_input_files("#file", arq); pg.wait_for_timeout(2600)
    pg.click("#toEquipe"); pg.wait_for_timeout(300)
    pg.click("#run"); pg.wait_for_timeout(2800)
    return pg.evaluate("""() => state.funil && {
      comparadas: state.funil.ufsComparadas,
      semRef: state.funil.ufsSemReferencia,
      naoConf: (state.funil.ufsNaoConfiaveis||[]).map(u=>u.uf),
      mesma: state.funil.ufsMesmaBase,
      conf: state.funil.confiavel,
      taxa: state.funil.taxaGeral,
      conv: state.funil.totalConversoes,
      cart: state.funil.totalCarteira,
      vend: state.funil.vendedores.length
    }""")


with sync_playwright() as p:
    b = p.chromium.launch(executable_path="/opt/pw-browsers/chromium")
    pg = b.new_context(viewport={"width": 1400, "height": 1000}).new_page()
    errs = []
    pg.on("pageerror", lambda e: errs.append(str(e)))
    pg.on("console", lambda m: errs.append(m.text) if m.type == "error" else None)
    pg.goto("file://" + APP)
    pg.wait_for_timeout(700)

    print("\n1. primeira rodada do PA — sem referência ainda")
    f = rodar(pg, f"{S}/pa_1.xlsx")
    ok(f["semRef"] == ["PA"], "PA marcado como sem referência", f)
    ok(not f["conf"], "sem aproveitamento a mostrar")

    print("\n2. primeira rodada do TO — não pode contaminar o PA")
    f = rodar(pg, f"{S}/to_1.xlsx")
    ok(f["semRef"] == ["TO"], "TO sem referência, PA nem entra na conta", f)
    ok(not f["naoConf"], "nenhum estado marcado como não confiável", f["naoConf"])
    refs = pg.evaluate("Object.keys(lerReferencias())")
    print("     referências guardadas:", refs)
    ok(sorted(refs) == ["normal|PA", "normal|TO"], "PA e TO convivem nas referências", refs)

    print("\n3. segunda rodada do PA — compara com o PA, não com o TO")
    f = rodar(pg, f"{S}/pa_2.xlsx")
    print("     ", f)
    ok(f["comparadas"] == ["PA"], "comparou PA", f["comparadas"])
    ok(f["conf"], "aproveitamento calculado")
    ok(0.05 < f["taxa"] < 0.30, f"taxa plausível ({f['taxa']*100:.1f}%)")
    ok(f["vend"] > 0, "ranking por vendedor voltou a aparecer", f["vend"])
    hist = pg.evaluate("lerHistorico()")
    ok(len(hist) == 1, "1 rodada no histórico", len(hist))
    ok([u["uf"] for u in hist[0]["ufs"]] == ["PA"], "histórico do PA", hist[0]["ufs"])

    print("\n4. segunda rodada do TO — a referência do PA continua intacta")
    f = rodar(pg, f"{S}/to_2.xlsx")
    print("     ", f)
    ok(f["comparadas"] == ["TO"], "comparou TO", f["comparadas"])
    ok(f["conf"], "aproveitamento calculado para o TO")
    hist = pg.evaluate("lerHistorico()")
    ok(len(hist) == 2, "2 rodadas no histórico", len(hist))

    print("\n5. histórico tem os dois estados")
    pg.click("#s4"); pg.wait_for_timeout(700)
    ufs = pg.evaluate("()=>[...document.querySelectorAll('#histUfs input')].map(i=>i.value)")
    print("     seletor:", ufs)
    ok("PA" in ufs and "TO" in ufs, "PA e TO no seletor", ufs)
    for uf in ["PA", "TO"]:
        pg.evaluate(f"()=>document.querySelector('#histUfs input[value={uf}]').click()")
        pg.wait_for_timeout(500)
        n = pg.evaluate("document.querySelectorAll('#histConteudo tbody tr').length")
        ok(n == 1, f"{uf} com 1 rodada no histórico", n)

    print("\n6. as análises por vendedor e por estado voltaram")
    ins = pg.evaluate("state.insights.map(i=>i.titulo)")
    for t in ins:
        print("     ", t)
    ok(any("Aproveitamento" in t for t in ins), "cartão de aproveitamento presente", ins)
    ok(pg.evaluate("document.querySelectorAll('#funil .frow').length") > 0,
       "tabela por vendedor renderizada")

    print("\n7. base com os sete estados de uma vez")
    f = rodar(pg, f"{S}/data_2.xlsx")
    print("     comparadas:", f["comparadas"], "| sem ref:", f["semRef"],
          "| não confiáveis:", f["naoConf"])
    ok(set(f["comparadas"]) >= {"PA", "TO"}, "PA e TO comparados na base completa", f["comparadas"])
    ok(set(f["semRef"]) == {"AC", "AM", "AP", "RO", "RR"},
       "os cinco estados novos entram agora", f["semRef"])
    pg.click("#s4"); pg.wait_for_timeout(700)
    sub = pg.inner_text("#perfSub")
    print("     texto:", sub[:190])
    ok("AC" in sub and "passa" in sub, "a tela explica quem entrou agora", sub[:120])
    pg.screenshot(path="estado-final.png", full_page=True)

    print("\nerros:", errs or "nenhum")
    if errs:
        falhas.append("erros de console")
    b.close()

print("\n" + (f"{len(falhas)} FALHA(S): {falhas}" if falhas else "Todos os testes passaram."))
raise SystemExit(1 if falhas else 0)
